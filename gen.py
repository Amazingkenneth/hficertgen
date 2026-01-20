import os
import threading
import datetime
import customtkinter as ctk
import tkinter.messagebox as messagebox  # Native dialogs still look best
from tkinter import filedialog
import pypinyin
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from tksheet import Sheet  # The best modern table widget for pasting/editing

# Configuration for CustomTkinter
ctk.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
ctk.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"


class CertificateGeneratorApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Window Setup
        self.title("Bilingual School Certificate Generator (Pro Version)")
        self.geometry("1200x800")

        # State Variables
        self.template_path = ctk.StringVar()
        self.excel_path = ctk.StringVar()
        self.output_dir = ctk.StringVar(value=os.path.join(os.getcwd(), "Output"))

        # Store data keys (headers) to map back to dictionary later
        self.data_keys = []

        self._setup_ui()

    def _setup_ui(self):
        # Configure Grid Layout
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)  # The table gets the most space

        # === 1. Configuration Area ===
        config_frame = ctk.CTkFrame(self, corner_radius=10)
        config_frame.grid(row=0, column=0, padx=20, pady=20, sticky="ew")
        config_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            config_frame, text="Configuration", font=("Roboto", 16, "bold")
        ).grid(row=0, column=0, padx=10, pady=(10, 5), sticky="w")

        # Template Selection
        ctk.CTkLabel(config_frame, text="Word Template:").grid(
            row=1, column=0, padx=10, pady=5, sticky="w"
        )
        ctk.CTkEntry(config_frame, textvariable=self.template_path).grid(
            row=1, column=1, padx=10, pady=5, sticky="ew"
        )
        ctk.CTkButton(
            config_frame, text="Browse", width=100, command=self.load_template
        ).grid(row=1, column=2, padx=10, pady=5)

        # Excel Selection
        ctk.CTkLabel(config_frame, text="Student Data:").grid(
            row=2, column=0, padx=10, pady=5, sticky="w"
        )
        ctk.CTkEntry(config_frame, textvariable=self.excel_path).grid(
            row=2, column=1, padx=10, pady=5, sticky="ew"
        )
        ctk.CTkButton(
            config_frame,
            text="Import Excel",
            width=100,
            command=self.load_excel,
            fg_color="#04A760",
            hover_color="#229965",
        ).grid(row=2, column=2, padx=10, pady=5)

        # Output Folder
        ctk.CTkLabel(config_frame, text="Output Folder:").grid(
            row=3, column=0, padx=10, pady=5, sticky="w"
        )
        ctk.CTkEntry(config_frame, textvariable=self.output_dir).grid(
            row=3, column=1, padx=10, pady=5, sticky="ew"
        )
        ctk.CTkButton(
            config_frame, text="Browse", width=100, command=self.select_output_dir
        ).grid(row=3, column=2, padx=10, pady=(5, 10))

        # === 2. Data Editor (tksheet) ===
        table_frame = ctk.CTkFrame(self, corner_radius=10)
        table_frame.grid(row=1, column=0, padx=20, pady=0, sticky="nsew")

        # Header for Table section
        header_frame = ctk.CTkFrame(table_frame, fg_color="transparent")
        header_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(
            header_frame, text="Data Editor", font=("Roboto", 16, "bold")
        ).pack(side="left")
        ctk.CTkLabel(
            header_frame,
            text="(Directly edit cells or Paste from Excel using Ctrl+V)",
            text_color="gray",
        ).pack(side="left", padx=10)

        # Initialize Sheet
        self.sheet = Sheet(
            table_frame,
            theme="dark blue" if ctk.get_appearance_mode() == "Dark" else "light blue",
            empty_horizontal=0,
            empty_vertical=0,
            font=("Roboto", 16, "normal"),
            header_font=("Roboto", 16, "bold"),
            index_font=("Roboto", 14, "normal"),
        )
        self.sheet.pack(fill="both", expand=True, padx=10, pady=10)

        # Enable Editing and Pasting bindings
        self.sheet.enable_bindings(
            (
                "single_select",
                "drag_select",
                "row_select",
                "column_select",
                "edit_cell",
                "paste",
                "cut",
                "copy",
                "delete",
                "arrowkeys",
                "rc_select",  # Right click select
                "rc_insert_row",  # Right click insert row
                "rc_delete_row",  # Right click delete row
            )
        )

        # === 3. Action Area ===
        action_frame = ctk.CTkFrame(self, corner_radius=10, height=80)
        action_frame.grid(row=2, column=0, padx=20, pady=20, sticky="ew")

        self.status_label = ctk.CTkLabel(action_frame, text="Ready", text_color="gray")
        self.status_label.pack(side="left", padx=20)

        self.generate_btn = ctk.CTkButton(
            action_frame,
            text="Generate Certificates",
            font=("Roboto", 15, "bold"),
            height=40,
            state="disabled",
            command=self.start_generation_thread,
        )
        self.generate_btn.pack(side="right", padx=20, pady=15)

        self.progress_bar = ctk.CTkProgressBar(action_frame, height=10)
        self.progress_bar.set(0)
        self.progress_bar.pack(side="right", fill="x", expand=True, padx=20)

    # ==========================
    # Logic Implementation
    # ==========================

    def load_template(self):
        f = filedialog.askopenfilename(filetypes=[("Word Documents", "*.docx")])
        if f:
            self.template_path.set(f)

    def select_output_dir(self):
        d = filedialog.askdirectory()
        if d:
            self.output_dir.set(d)

    def load_excel(self):
        filename = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx *.xlsm")]
        )
        if not filename:
            return
        self.excel_path.set(filename)
        self.status_label.configure(text="Processing Excel data...")
        threading.Thread(
            target=self._process_excel_thread, args=(filename,), daemon=True
        ).start()

    def _process_excel_thread(self, filename):
        try:
            wb = load_workbook(filename, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            # Map columns
            col_indices = self._map_columns(headers)
            if not col_indices:
                self.after(
                    0,
                    lambda: messagebox.showerror(
                        "Error", "Could not find required columns (Name, ID, etc)."
                    ),
                )
                return

            processed_rows = []

            # Use iter_rows to read data
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Basic validation
                name_idx = col_indices.get("name_zh")
                if not row[name_idx]:
                    continue

                entry_dict = self._parse_row_data(row, col_indices)
                processed_rows.append(entry_dict)

            # Update UI in main thread
            self.after(0, lambda: self._update_table_data(processed_rows))

        except Exception as e:
            self.after(
                0,
                lambda: messagebox.showerror(
                    "Error", f"Failed to load Excel: {str(e)}"
                ),
            )

    def _update_table_data(self, data_list):
        if not data_list:
            self.status_label.configure(text="No valid data found.")
            return

        # Extract headers (keys) from the first dictionary
        self.data_keys = list(data_list[0].keys())

        # Transform list of dicts -> list of lists for tksheet
        sheet_data = []
        for entry in data_list:
            row_data = [entry.get(k, "") for k in self.data_keys]
            sheet_data.append(row_data)

        # Set headers and data
        self.sheet.headers(self.data_keys)
        self.sheet.set_sheet_data(sheet_data)

        # Auto resize columns to fit data
        self.sheet.column_width(column="all", width="text")

        self.generate_btn.configure(state="normal")
        self.status_label.configure(
            text=f"Loaded {len(data_list)} students. You can now edit data above."
        )

    def start_generation_thread(self):
        if not self.template_path.get():
            messagebox.showerror("Error", "Please select a template file.")
            return

        # Lock the button
        self.generate_btn.configure(state="disabled", text="Generating...")
        threading.Thread(target=self.generate_docs, daemon=True).start()

    def generate_docs(self):
        tpl_path = self.template_path.get()
        out_path = self.output_dir.get()
        if not os.path.exists(out_path):
            os.makedirs(out_path)

        # Get data directly from the sheet (allows for manual edits)
        current_data = self.sheet.get_sheet_data()
        total = len(current_data)

        # Reset progress
        self.after(0, lambda: self.progress_bar.set(0))

        try:
            doc = DocxTemplate(tpl_path)

            for i, row_values in enumerate(current_data):
                # Reconstruct dictionary context from sheet row
                context = dict(zip(self.data_keys, row_values))

                # Update progress
                progress = (i + 1) / total
                self.after(
                    0, lambda p=progress, idx=i: self._update_progress_ui(p, idx, total)
                )

                # Render
                doc.render(context)

                safe_name = str(context.get("name_zh", "Doc")).replace("/", "_")
                safe_id = str(context.get("student_id", "Unknown"))
                filename = f"{safe_id}_{safe_name}.docx"
                doc.save(os.path.join(out_path, filename))

            self.after(0, self._finish_generation_ui)

        except Exception as e:
            self.after(
                0, lambda: messagebox.showerror("Error", f"Generation failed: {str(e)}")
            )
            self.after(0, self._reset_ui)

    def _update_progress_ui(self, progress, current, total):
        self.progress_bar.set(progress)
        self.status_label.configure(text=f"Generating {current + 1}/{total}...")

    def _finish_generation_ui(self):
        self.progress_bar.set(1.0)
        self.status_label.configure(text="Completed!")
        self.generate_btn.configure(state="normal", text="Generate Certificates")
        messagebox.showinfo("Success", f"Files generated in:\n{self.output_dir.get()}")

    def _reset_ui(self):
        self.generate_btn.configure(state="normal", text="Generate Certificates")

    # ==========================
    # Helper Logic (Kept similar)
    # ==========================

    def _map_columns(self, headers):
        mapping = {}
        keywords = {
            "name_zh": ["姓名", "legal name", "Name"],
            "gender_zh": ["性别", "Gender"],
            "id_type_zh": ["证件类型", "ID Type"],
            "id_number": ["身份证件号码", "No.", "ID Number"],
            "dob": ["出生日期", "Birth"],
            "admit_date": ["入学年份", "Admission"],
            "student_id": ["学号", "Student ID"],
            "grade": ["在读年级", "Grade"],
        }
        headers_str = [str(h) if h else "" for h in headers]
        for key, kw_list in keywords.items():
            found = False
            for idx, header in enumerate(headers_str):
                for kw in kw_list:
                    if kw in header:
                        mapping[key] = idx
                        found = True
                        break
                if found:
                    break
            if not found:
                mapping[key] = -1

        if mapping.get("name_zh", -1) == -1:
            return None
        return mapping

    def _parse_row_data(self, row, idx_map):
        def get_val(key):
            idx = idx_map.get(key, -1)
            if idx == -1 or idx >= len(row):
                return ""
            return str(row[idx]).strip() if row[idx] is not None else ""

        entry = {}
        current_date = datetime.date.today()

        # Basic fields
        entry["name_zh"] = get_val("name_zh")
        entry["gender_zh"] = get_val("gender_zh")
        entry["id_type_zh"] = get_val("id_type_zh") or "身份证"
        entry["grade"] = get_val("grade")
        entry["student_id"] = get_val("student_id")

        # ID cleanup
        raw_id = get_val("id_number")
        entry["id_number"] = (
            raw_id.replace("‘", "").replace("’", "").replace("'", "").strip()
        )

        # Date Parsing
        raw_dob = row[idx_map.get("dob", -1)] if idx_map.get("dob", -1) != -1 else ""
        dob_dt = self._smart_date_parse(raw_dob)
        if dob_dt:
            entry["dob_zh"] = dob_dt.strftime("%Y年%m月%d日")
            entry["dob_en"] = dob_dt.strftime("%B %d, %Y")
        else:
            entry["dob_zh"] = str(raw_dob)
            entry["dob_en"] = str(raw_dob)

        # Admit Date
        raw_admit = (
            row[idx_map.get("admit_date", -1)]
            if idx_map.get("admit_date", -1) != -1
            else ""
        )
        admit_dt = self._smart_date_parse(raw_admit)
        if admit_dt:
            entry["admit_year"] = str(admit_dt.year)
            entry["admit_month"] = str(admit_dt.month)
            entry["admit_month_en"] = admit_dt.strftime("%B")
        else:
            # Fallback for text like "2024-08"
            if isinstance(raw_admit, str) and "-" in raw_admit:
                parts = raw_admit.split("-")
                entry["admit_year"] = parts[0]
                entry["admit_month"] = parts[1] if len(parts) > 1 else "??"
                import calendar

                try:
                    entry["admit_month_en"] = calendar.month_name[int(parts[1])]
                except:
                    entry["admit_month_en"] = "Unknown"
            else:
                entry["admit_year"] = entry["admit_month"] = entry["admit_month_en"] = (
                    "Unknown"
                )

        # Current Date
        entry["date_zh"] = current_date.strftime("%Y年%m月%d日")
        entry["date_en"] = current_date.strftime("%B %d, %Y")

        # Translations
        entry["name_en"] = get_english_name(entry["name_zh"])

        # Gender Translation
        if "男" in entry["gender_zh"]:
            entry["gender_en"] = "Male"
        elif "女" in entry["gender_zh"]:
            entry["gender_en"] = "Female"
        else:
            entry["gender_en"] = entry["gender_zh"]

        # ID Type Translation
        if "身份证" in entry["id_type_zh"]:
            entry["id_type_en"] = "ID Card"
        elif "护照" in entry["id_type_zh"]:
            entry["id_type_en"] = "Passport"
        else:
            entry["id_type_en"] = "ID Document"

        return entry

    def _smart_date_parse(self, val):
        if val is None:
            return None
        if isinstance(val, (datetime.datetime, datetime.date)):
            return val
        val_str = str(val).strip()
        if not val_str:
            return None

        formats = [
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d",
            "%m/%d/%Y",
            "%Y-%m-%d",
            "%Y-%m",
            "%Y.%m.%d",
            "%Y%m%d",
        ]
        for fmt in formats:
            try:
                return datetime.datetime.strptime(val_str, fmt)
            except ValueError:
                continue
        return None


def get_english_name(chinese_name):
    # (Helper function logic same as before)
    if all(ord(c) < 128 for c in chinese_name if c.isalpha()):
        return chinese_name
    if not chinese_name or len(chinese_name) < 2:
        return ""
    py_list = pypinyin.pinyin(chinese_name, style=pypinyin.Style.NORMAL)
    if not py_list:
        return ""
    surname = py_list[0][0].capitalize()
    firstname = "".join([x[0] for x in py_list[1:]]).capitalize()
    return f"{surname}, {firstname}"


if __name__ == "__main__":
    app = CertificateGeneratorApp()
    app.mainloop()
